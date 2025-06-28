import tkinter as tk
from tkinter import messagebox, font
from threading import Thread
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from datetime import datetime

#Config para manter o navegador aberto ao finalizar a execução
#options = Options()
#options.add_experimental_option("detach", True)

def executar_coleta():

    navegador = webdriver.Chrome()

    #Abre o navegador e pesquisa o site de clima
    navegador.get('https://www.climatempo.com.br/previsao-do-tempo/agora/cidade/558/saopaulo-sp')

    # Pegar temperatura
    temperatura = navegador.find_element(By.XPATH, '//*[@id="mainContent"]/div[6]/div[3]/div[1]/div[3]/div[1]/div/div[2]/span').text

    # Pegar sensação
    sensacao = navegador.find_element(By.XPATH, '//*[@id="mainContent"]/div[6]/div[3]/div[1]/div[3]/div[1]/div/div[3]/span').text

    # Pegar umidade
    umidade = navegador.find_element(By.XPATH, '//*[@id="mainContent"]/div[6]/div[3]/div[1]/div[3]/div[1]/div/div[4]/ul/li[2]/div[2]/p/span').text

    agora = datetime.now()
    data = agora.strftime("%d/%m/%Y")
    hora = agora.strftime("%H:%M:%S")

    #Abre o execel
    arquivo = load_workbook('temperaturas-salvas.xlsx')

    plan = arquivo['Clima']

    # Descobre a próxima linha vazia
    ultima_linha = plan.max_row + 1

    # Preenche os dados na próxima linha
    plan.cell(row=ultima_linha, column=1).value = temperatura
    plan.cell(row=ultima_linha, column=2).value = sensacao
    plan.cell(row=ultima_linha, column=3).value = umidade
    plan.cell(row=ultima_linha, column=4).value = data
    plan.cell(row=ultima_linha, column=5).value = hora

    arquivo.save('temperaturas-salvas.xlsx')
    print('Planilha atualizada com sucesso!')

    navegador.close()

def iniciar_thread():
    Thread(target=executar_coleta, daemon=True).start()

root = tk.Tk()
root.title("🌤️ Capturador de Temperatura")
root.geometry("350x250")
root.configure(bg="#87CEEB")  # Azul céu claro

# Fonte personalizada
fonte_titulo = font.Font(family="Helvetica", size=16, weight="bold")
fonte_texto = font.Font(family="Arial", size=12)

# Título
lbl_titulo = tk.Label(root, text="🌤️ Capturador de Temperatura", bg="#87CEEB", fg="#ffffff", font=fonte_titulo)
lbl_titulo.pack(pady=10)

# Botão pesquisar
btn_pesquisar = tk.Button(root, text="Pesquisar a temperatura atual", font=fonte_texto, bg="#ffcc00", fg="#000000",
                          activebackground="#ffaa00", padx=10, pady=5, command=iniciar_thread)
btn_pesquisar.pack(pady=10)

# Label para mostrar resultado
lbl_resultado = tk.Label(root, text="", bg="#87CEEB", fg="#ffffff", font=fonte_texto, justify="left")
lbl_resultado.pack(pady=10)

root.mainloop()
